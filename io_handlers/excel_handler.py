"""
Handler per l'import/export in formato Excel (.xlsx).
Struttura Excel:
- Sheet "Metadati": informazioni azienda e assessor
- Sheet "Risposte": tutte le risposte con modulo, domanda, valore, note
- Sheet "Risultati": punteggi per modulo
- Sheet "Roadmap": azioni pianificate
- Sheet "Domande Custom": domande aggiunte dall'utente
"""
from __future__ import annotations
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models.schema import AssessmentData
from models.questions import ALL_QUESTIONS, QUESTIONS_BY_ID, get_questions_for_module
from config import MODULES, MODULE_MAP


# ── Stili Excel ─────────────────────────────────────────────────────────────
_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="3B3689", end_color="3B3689", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row: int, max_col: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def save_to_excel(data: AssessmentData) -> bytes:
    """Esporta l'assessment completo in un file Excel."""
    wb = Workbook()

    # ── Sheet 1: Metadati ───────────────────────────────────────────────
    ws_meta = wb.active
    ws_meta.title = "Metadati"
    meta_fields = [
        ("Campo", "Valore"),
        ("Nome azienda", data.company_name),
        ("Settore", data.company_sector),
        ("Dimensione", data.company_size),
        ("Fatturato", data.company_revenue),
        ("N. dipendenti", data.company_employees),
        ("Descrizione", data.company_description),
        ("Assessor", data.assessor_name),
        ("Ruolo assessor", data.assessor_role),
        ("Creato il", data.created_at),
        ("Aggiornato il", data.updated_at),
        ("Versione", data.version),
    ]
    for r, (campo, valore) in enumerate(meta_fields, 1):
        ws_meta.cell(row=r, column=1, value=campo)
        ws_meta.cell(row=r, column=2, value=str(valore) if valore else "")
    _style_header(ws_meta, 1, 2)
    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 50

    # ── Sheet 2: Risposte ──────────────────────────────────────────────
    ws_ans = wb.create_sheet("Risposte")
    headers = ["ID Domanda", "Modulo", "Rif. ISO", "Tipo", "Domanda", "Valore", "Note"]
    for c, h in enumerate(headers, 1):
        ws_ans.cell(row=1, column=c, value=h)
    _style_header(ws_ans, 1, len(headers))

    row = 2
    for mod in MODULES:
        qs = get_questions_for_module(mod["id"], data.custom_questions)
        for q in qs:
            ans = data.get_answer(q["id"])
            val = ans["value"] if ans else ""
            notes = ans.get("notes", "") if ans else ""
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)

            ws_ans.cell(row=row, column=1, value=q["id"])
            ws_ans.cell(row=row, column=2, value=mod["name"])
            ws_ans.cell(row=row, column=3, value=q.get("iso_ref", ""))
            ws_ans.cell(row=row, column=4, value=q["type"])
            ws_ans.cell(row=row, column=5, value=q["text"])
            ws_ans.cell(row=row, column=6, value=str(val) if val else "")
            ws_ans.cell(row=row, column=7, value=notes)
            row += 1

    ws_ans.column_dimensions["A"].width = 14
    ws_ans.column_dimensions["B"].width = 30
    ws_ans.column_dimensions["C"].width = 10
    ws_ans.column_dimensions["D"].width = 14
    ws_ans.column_dimensions["E"].width = 60
    ws_ans.column_dimensions["F"].width = 30
    ws_ans.column_dimensions["G"].width = 40

    # ── Sheet 3: Risultati ─────────────────────────────────────────────
    ws_res = wb.create_sheet("Risultati")
    res_headers = ["Modulo", "Rif. ISO", "Punteggio", "Livello maturità", "Target", "Gap", "Completamento"]
    for c, h in enumerate(res_headers, 1):
        ws_res.cell(row=1, column=c, value=h)
    _style_header(ws_res, 1, len(res_headers))

    row = 2
    for mod in MODULES:
        mr = data.module_results.get(mod["id"], {})
        ws_res.cell(row=row, column=1, value=mod["name"])
        ws_res.cell(row=row, column=2, value=mod["iso_ref"])
        ws_res.cell(row=row, column=3, value=mr.get("score", 0))
        ws_res.cell(row=row, column=4, value=mr.get("maturity_level", 0))
        ws_res.cell(row=row, column=5, value=mr.get("target_level", 3))
        ws_res.cell(row=row, column=6, value=mr.get("gap", 0))
        ws_res.cell(row=row, column=7, value=f"{mr.get('answered', 0)}/{mr.get('total', 0)}")
        row += 1

    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws_res.column_dimensions[col_letter].width = 20
    ws_res.column_dimensions["A"].width = 35

    # ── Sheet 4: Roadmap ───────────────────────────────────────────────
    ws_road = wb.create_sheet("Roadmap")
    road_headers = ["ID", "Titolo", "Descrizione", "Modulo", "Priorità", "Effort", "Impatto", "Timeframe", "Owner", "Stato", "Note"]
    for c, h in enumerate(road_headers, 1):
        ws_road.cell(row=1, column=c, value=h)
    _style_header(ws_road, 1, len(road_headers))

    for r, item in enumerate(data.roadmap, 2):
        if isinstance(item, dict):
            ws_road.cell(row=r, column=1, value=item.get("id", ""))
            ws_road.cell(row=r, column=2, value=item.get("title", ""))
            ws_road.cell(row=r, column=3, value=item.get("description", ""))
            mod_name = MODULE_MAP.get(item.get("module_id", ""), {}).get("name", item.get("module_id", ""))
            ws_road.cell(row=r, column=4, value=mod_name)
            ws_road.cell(row=r, column=5, value=item.get("priority", ""))
            ws_road.cell(row=r, column=6, value=item.get("effort", ""))
            ws_road.cell(row=r, column=7, value=item.get("impact", ""))
            ws_road.cell(row=r, column=8, value=item.get("timeframe", ""))
            ws_road.cell(row=r, column=9, value=item.get("owner", ""))
            ws_road.cell(row=r, column=10, value=item.get("status", ""))
            ws_road.cell(row=r, column=11, value=item.get("notes", ""))

    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
        ws_road.column_dimensions[col_letter].width = 18
    ws_road.column_dimensions["C"].width = 50

    # ── Sheet 5: Domande Custom ────────────────────────────────────────
    ws_custom = wb.create_sheet("Domande Custom")
    custom_headers = ["ID", "Modulo", "Tipo", "Testo", "Opzioni", "Peso", "Tip"]
    for c, h in enumerate(custom_headers, 1):
        ws_custom.cell(row=1, column=c, value=h)
    _style_header(ws_custom, 1, len(custom_headers))

    for r, q in enumerate(data.custom_questions, 2):
        ws_custom.cell(row=r, column=1, value=q.get("id", ""))
        ws_custom.cell(row=r, column=2, value=q.get("module", ""))
        ws_custom.cell(row=r, column=3, value=q.get("type", ""))
        ws_custom.cell(row=r, column=4, value=q.get("text", ""))
        ws_custom.cell(row=r, column=5, value=", ".join(q.get("options", [])))
        ws_custom.cell(row=r, column=6, value=q.get("weight", 1.0))
        ws_custom.cell(row=r, column=7, value=q.get("tip", ""))

    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws_custom.column_dimensions[col_letter].width = 18
    ws_custom.column_dimensions["D"].width = 60

    # ── Salva in bytes ─────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def load_from_excel(file_obj) -> AssessmentData:
    """Importa un assessment da file Excel."""
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    data = AssessmentData()

    # ── Metadati ───────────────────────────────────────────────────────
    if "Metadati" in wb.sheetnames:
        ws = wb["Metadati"]
        meta_map = {}
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] and row[1]:
                meta_map[row[0]] = str(row[1])

        data.company_name = meta_map.get("Nome azienda", "")
        data.company_sector = meta_map.get("Settore", "")
        data.company_size = meta_map.get("Dimensione", "")
        data.company_revenue = meta_map.get("Fatturato", "")
        data.company_employees = meta_map.get("N. dipendenti", "")
        data.company_description = meta_map.get("Descrizione", "")
        data.assessor_name = meta_map.get("Assessor", "")
        data.assessor_role = meta_map.get("Ruolo assessor", "")

    # ── Risposte ───────────────────────────────────────────────────────
    if "Risposte" in wb.sheetnames:
        ws = wb["Risposte"]
        for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            qid = row[0]
            val = row[5]
            notes = row[6] or ""
            if qid and val is not None and str(val).strip() != "":
                # Ricostruisci il tipo dalla domanda
                q_def = QUESTIONS_BY_ID.get(qid)
                if q_def and q_def["type"] == "multi_choice" and isinstance(val, str):
                    val = [v.strip() for v in val.split(",") if v.strip()]
                elif q_def and q_def["type"] == "scale":
                    try:
                        val = int(float(val))
                    except (ValueError, TypeError):
                        pass
                data.set_answer(qid, val, str(notes))

    # ── Domande Custom ─────────────────────────────────────────────────
    if "Domande Custom" in wb.sheetnames:
        ws = wb["Domande Custom"]
        for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            if row[0] and row[3]:
                q = {
                    "id": str(row[0]),
                    "module": str(row[1] or ""),
                    "type": str(row[2] or "scale"),
                    "text": str(row[3]),
                    "options": [o.strip() for o in str(row[4] or "").split(",") if o.strip()],
                    "weight": float(row[5]) if row[5] else 1.0,
                    "tip": str(row[6] or ""),
                    "required": False,
                    "iso_ref": "",
                }
                data.custom_questions.append(q)

    # ── Target levels ──────────────────────────────────────────────────
    if "Risultati" in wb.sheetnames:
        ws = wb["Risultati"]
        for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
            if row[0] and row[4]:
                # Trova module_id dal nome
                for mod in MODULES:
                    if mod["name"] == row[0]:
                        try:
                            data.target_levels[mod["id"]] = int(float(row[4]))
                        except (ValueError, TypeError):
                            pass

    wb.close()
    data.touch()
    return data
